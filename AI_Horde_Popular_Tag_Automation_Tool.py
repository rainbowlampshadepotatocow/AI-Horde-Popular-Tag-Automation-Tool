import os
import requests
import json
import pandas as pd

# Constants
API_URL = "https://aihorde.net/api/v2/stats/text/models"

# Directory where input and output files are stored
USER_FILES_DIR = os.path.join(os.path.dirname(__file__), "user-files")
os.makedirs(USER_FILES_DIR, exist_ok=True)

USAGE_JSON = os.path.join(USER_FILES_DIR, "RawModelUsageData.json")
MODELS_CSV = os.path.join(USER_FILES_DIR, "models.csv")
OUTPUT_CSV = os.path.join(USER_FILES_DIR, "models_updated.csv")
TOP_N = 25
PERIOD_TAGS = {
    'day': 'daily_top25',
    'month': 'monthly_top25',
    'total': 'alltime_top25'
}

# Step 1: Fetch latest usage data
print(f"Fetching usage data from {API_URL}...")
r = requests.get(API_URL)
r.raise_for_status()
usage_data = r.json()

# Optionally save raw JSON
with open(USAGE_JSON, 'w', encoding='utf-8') as f:
    json.dump(usage_data, f, indent=2)

# Step 2: Build DataFrame of usage
records = []
for period, models in usage_data.items():
    for model_name, count in models.items():
        records.append({'period': period, 'model': model_name, 'usage_count': count})

df_usage = pd.DataFrame(records)
df_usage.to_csv(os.path.join(USER_FILES_DIR, "usage_data.csv"), index=False)

# Convert the usage data to an Excel workbook with one sheet per period
usage_xlsx = os.path.join(USER_FILES_DIR, "usage_data.xlsx")
with pd.ExcelWriter(usage_xlsx) as writer:
    for period, group in df_usage.groupby("period"):
        sheet = period.capitalize()
        # Include headers when writing each sheet
        group.drop(columns="period").to_excel(writer, index=False, sheet_name=sheet)

# Format each sheet as a table once written
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook(usage_xlsx)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    max_row = ws.max_row
    max_col = ws.max_column
    end_col = get_column_letter(max_col)
    table_ref = f"A1:{end_col}{max_row}"
    table = Table(displayName=f"{sheet}Table", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

wb.save(usage_xlsx)

# Step 3: Determine top N models per period
top_models = {}
for period, group in df_usage.groupby('period'):
    # Sort by usage and take top N
    top_list = group.sort_values('usage_count', ascending=False)
    top_models[period] = set(top_list.head(TOP_N)['model'].tolist())
    print(f"Top {TOP_N} for {period}: {len(top_models[period])} models")


## STOP HERE FOR NOW ##
# # Step 4: Load existing models whitelist
# df_models = pd.read_csv(MODELS_CSV)
# # Ensure tags column exists
# if 'tags' not in df_models.columns:
#     df_models['tags'] = ''

# # Step 5: Update tags based on membership in top lists
# new_tags = []
# for _, row in df_models.iterrows():
#     model = row['name']
#     tags = []
#     for period, tag in PERIOD_TAGS.items():
#         if model in top_models.get(period, set()):
#             tags.append(tag)
#     # Join with commas, preserve any existing tags if desired
#     new_tags.append(','.join(tags))

# df_models['tags'] = new_tags

# # Step 6: Save updated CSV
# df_models.to_csv(OUTPUT_CSV, index=False)
# print(f"Updated CSV written to {OUTPUT_CSV}")
